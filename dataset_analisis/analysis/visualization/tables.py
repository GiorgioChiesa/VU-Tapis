"""
Table Export Module for GraSP Analysis.
Exports analysis results to CSV and Excel formats.
"""

import logging
import os
from typing import Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from config.analysis_config import OUTPUT_CONFIG


class TableExporter:
    """
    Exports analysis results to CSV and Excel formats.
    """

    def __init__(self, logger=None):
        """Initialize table exporter."""
        self.logger = logger or logging.getLogger(__name__)
        self.output_dir = OUTPUT_CONFIG['tables_dir']
        os.makedirs(self.output_dir, exist_ok=True)

    def export_step_distribution(self, df: pd.DataFrame):
        """Export step distribution table."""
        self.logger.info("Exporting step distribution table...")
        output_path = os.path.join(self.output_dir, '01_step_distribution.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_step_video_matrix(self, df: pd.DataFrame):
        """Export step × video matrix."""
        self.logger.info("Exporting step-video matrix...")
        output_path = os.path.join(self.output_dir, '02_step_video_matrix.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_video_statistics(self, df: pd.DataFrame):
        """Export video statistics."""
        self.logger.info("Exporting video statistics...")
        output_path = os.path.join(self.output_dir, '03_video_statistics.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_step_phase_mapping(self, df: pd.DataFrame):
        """Export step-phase mapping."""
        self.logger.info("Exporting step-phase mapping...")
        output_path = os.path.join(self.output_dir, '04_step_phase_mapping.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_temporal_progression(self, df: pd.DataFrame):
        """Export temporal progression data."""
        self.logger.info("Exporting temporal progression...")
        output_path = os.path.join(self.output_dir, '05_temporal_progression.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_step_transitions(self, df: pd.DataFrame):
        """Export step transitions."""
        self.logger.info("Exporting step transitions...")
        output_path = os.path.join(self.output_dir, '06_step_transitions.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_train_test_comparison(self, df: pd.DataFrame):
        """Export train-test comparison."""
        self.logger.info("Exporting train-test comparison...")
        output_path = os.path.join(self.output_dir, '07_train_test_comparison.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_imbalance_metrics(self, imbalance_dict: dict, imbalance_summary_df: pd.DataFrame = None):
        """Export imbalance metrics."""
        self.logger.info("Exporting imbalance metrics...")

        # Create DataFrame from dict
        rows = []
        for key, value in imbalance_dict.items():
            rows.append({'metric': key, 'value': value})

        df = pd.DataFrame(rows)
        output_path = os.path.join(self.output_dir, '08_imbalance_metrics.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_phase_distribution(self, df: pd.DataFrame):
        """Export phase distribution table."""
        self.logger.info("Exporting phase distribution table...")
        output_path = os.path.join(self.output_dir, '09_phase_distribution.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_phase_video_matrix(self, df: pd.DataFrame):
        """Export phase × video matrix."""
        self.logger.info("Exporting phase-video matrix...")
        output_path = os.path.join(self.output_dir, '10_phase_video_matrix.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_phase_composition(self, df: pd.DataFrame):
        """Export phase composition."""
        self.logger.info("Exporting phase composition...")
        output_path = os.path.join(self.output_dir, '11_phase_composition.csv')
        df.to_csv(output_path, index=False)
        self.logger.info(f"Saved: {output_path}")

    def export_to_excel(self, analysis_results: Dict):
        """
        Export all tables to a single Excel workbook with multiple sheets.

        Args:
            analysis_results: Dictionary with all analysis results
        """
        self.logger.info("Exporting all results to Excel workbook...")

        output_path = os.path.join(self.output_dir, 'MASTER_ANALYSIS.xlsx')

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Map of sheet names and corresponding data
        sheets_map = {
            'Step Distribution': analysis_results.get('step_distribution'),
            'Phase Distribution': analysis_results.get('phase_distribution'),
            'Step-Video Matrix': analysis_results.get('step_video_matrix'),
            'Phase-Video Matrix': analysis_results.get('phase_video_matrix'),
            'Video Statistics': analysis_results.get('video_statistics'),
            'Step-Phase Mapping': analysis_results.get('step_phase_correlation'),
            'Phase Composition': analysis_results.get('phase_composition'),
            'Temporal Progression': analysis_results.get('temporal_progression'),
            'Step Transitions': analysis_results.get('step_transitions'),
            'Train-Test Comparison': analysis_results.get('train_test_comparison'),
        }

        # Add data sheets
        for sheet_name, df in sheets_map.items():
            if df is not None and len(df) > 0:
                ws = wb.create_sheet(title=sheet_name)

                # Add data
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                self.logger.info(f"Added sheet: {sheet_name}")

        # Add summary metrics sheet
        ws_summary = wb.create_sheet(title='Summary')
        summary_data = self._create_summary_metrics(analysis_results)
        for r_idx, row in enumerate(dataframe_to_rows(summary_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        # Save workbook
        wb.save(output_path)
        self.logger.info(f"Saved Excel workbook: {output_path}")

    def _create_summary_metrics(self, analysis_results: Dict) -> pd.DataFrame:
        """Create summary metrics dataframe."""
        rows = []

        # Dataset sizes
        metrics = {
            'total_train_annotations': len(analysis_results.get('step_distribution', pd.DataFrame())) * 0,  # Placeholder
            'total_test_annotations': 0,  # Placeholder
            'total_steps': len(analysis_results.get('step_distribution', [])),
            'total_phases': len(analysis_results.get('phase_distribution', [])),
        }

        # Coverage metrics
        if 'step_coverage' in analysis_results:
            coverage = analysis_results['step_coverage']
            metrics.update({
                'steps_in_train': coverage.get('steps_in_train', 0),
                'steps_in_test': coverage.get('steps_in_test', 0),
                'steps_in_both': coverage.get('steps_in_both', 0),
                'coverage_train_pct': coverage.get('coverage_train', 0) * 100,
                'coverage_test_pct': coverage.get('coverage_test', 0) * 100,
            })

        # Imbalance metrics
        if 'imbalance_metrics' in analysis_results:
            imbalance = analysis_results['imbalance_metrics']
            metrics.update({
                'train_min_count': imbalance.get('train_min', 0),
                'train_max_count': imbalance.get('train_max', 0),
                'train_imbalance_ratio': imbalance.get('train_ratio', 0),
                'test_min_count': imbalance.get('test_min', 0),
                'test_max_count': imbalance.get('test_max', 0),
                'test_imbalance_ratio': imbalance.get('test_ratio', 0),
            })

        # Chi-square test
        if 'chi_square_test' in analysis_results:
            chi2_test = analysis_results['chi_square_test']
            metrics.update({
                'chi_square_statistic': chi2_test.get('chi2_statistic', 0),
                'chi_square_p_value': chi2_test.get('p_value', 0),
                'chi_square_significant': chi2_test.get('significant', False),
            })

        # Convert to rows
        for key, value in metrics.items():
            rows.append({'metric': key, 'value': value})

        return pd.DataFrame(rows)

    def export_all(self, analysis_results: Dict):
        """
        Export all analysis results to CSV and Excel.

        Args:
            analysis_results: Dictionary with all analysis results
        """
        self.logger.info("=" * 70)
        self.logger.info("EXPORTING ALL ANALYSIS RESULTS")
        self.logger.info("=" * 70)

        try:
            # Export CSV files
            if 'step_distribution' in analysis_results:
                self.export_step_distribution(analysis_results['step_distribution'])
            if 'phase_distribution' in analysis_results:
                self.export_phase_distribution(analysis_results['phase_distribution'])
            if 'step_video_matrix' in analysis_results:
                self.export_step_video_matrix(analysis_results['step_video_matrix'])
            if 'phase_video_matrix' in analysis_results:
                self.export_phase_video_matrix(analysis_results['phase_video_matrix'])
            if 'video_statistics' in analysis_results:
                self.export_video_statistics(analysis_results['video_statistics'])
            if 'step_phase_correlation' in analysis_results:
                self.export_step_phase_mapping(analysis_results['step_phase_correlation'])
            if 'phase_composition' in analysis_results:
                self.export_phase_composition(analysis_results['phase_composition'])
            if 'temporal_progression' in analysis_results:
                self.export_temporal_progression(analysis_results['temporal_progression'])
            if 'step_transitions' in analysis_results:
                self.export_step_transitions(analysis_results['step_transitions'])
            if 'train_test_comparison' in analysis_results:
                self.export_train_test_comparison(analysis_results['train_test_comparison'])
            if 'imbalance_metrics' in analysis_results:
                self.export_imbalance_metrics(analysis_results['imbalance_metrics'])

            # Export Excel
            self.export_to_excel(analysis_results)

            self.logger.info("=" * 70)
            self.logger.info("ALL TABLES EXPORTED SUCCESSFULLY")
            self.logger.info("=" * 70)

        except Exception as e:
            self.logger.error(f"Error exporting tables: {e}")
            raise
